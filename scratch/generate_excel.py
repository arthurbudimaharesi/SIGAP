import pandas as pd
from openpyxl import load_workbook
import shutil
import os

os.makedirs('docs', exist_ok=True)
template_path = 'docs/Test Report RPL (1).xlsx'
output_path = 'docs/Test Report - Arthur.xlsx'
shutil.copy(template_path, output_path)

wb = load_workbook(output_path)
ws = wb.active

test_cases_data = [
    {
        'pbi': 'PBI-01', 'case_id': 'TC-01-01', 'scenario': 'Administrasi Data Pelanggan', 'type': 'Positive', 
        'test_case': 'CRUD Data Pelanggan via Web', 'pre_cond': 'Akun admin',
        'steps': [
            (1, 'Launch the app dan Login sebagai admin', 'Masuk ke Dashboard Admin.', 'Pass'),
            (2, 'Arahkan ke menu Pelanggan', 'Tabel daftar pelanggan tampil.', 'Pass'),
            (3, 'Klik tombol Tambah Pelanggan', 'Form tambah pelanggan tampil.', 'Pass')
        ],
        'ket': '', 'evidence': 'SS_PBI-01_Pelanggan_Index.png'
    },
    {
        'pbi': 'PBI-02', 'case_id': 'TC-02-01', 'scenario': 'Administrasi Kategori Pengaduan', 'type': 'Positive', 
        'test_case': 'CRUD Data Kategori via Web', 'pre_cond': 'Akun admin',
        'steps': [
            (1, 'Login sebagai admin', 'Masuk ke Dashboard Admin.', 'Pass'),
            (2, 'Arahkan ke menu Kategori Pengaduan', 'Tabel daftar kategori tampil.', 'Pass'),
            (3, 'Klik tombol Tambah Kategori', 'Form tambah kategori tampil.', 'Pass')
        ],
        'ket': '', 'evidence': 'SS_PBI-02_Kategori_Index.png'
    },
    {
        'pbi': 'PBI-03 / PBI-20', 'case_id': 'TC-03-01', 'scenario': 'CRUD Zona Wilayah', 'type': 'Positive', 
        'test_case': 'Mengelola zona wilayah via Web', 'pre_cond': 'Akun admin',
        'steps': [
            (1, 'Login sebagai admin', 'Masuk ke Dashboard Admin.', 'Pass'),
            (2, 'Arahkan ke menu Zona Wilayah', 'Tabel daftar zona wilayah tampil.', 'Pass'),
            (3, 'Buka form Tambah Zona', 'Peta dan form zona wilayah tampil.', 'Pass')
        ],
        'ket': '', 'evidence': 'SS_PBI-03_PBI-20_Zona_Index.png'
    },
    {
        'pbi': 'PBI-21', 'case_id': 'TC-21-01', 'scenario': 'Pengelolaan Zona Wilayah Pengaduan', 'type': 'Positive', 
        'test_case': 'Supervisor memantau zona via Web', 'pre_cond': 'Akun supervisor',
        'steps': [
            (1, 'Login sebagai supervisor', 'Masuk ke Dashboard Supervisor.', 'Pass'),
            (2, 'Arahkan ke menu Zona', 'Daftar pemantauan zona wilayah tampil.', 'Pass')
        ],
        'ket': '', 'evidence': 'SS_PBI-21_Supervisor_Zona.png'
    },
    {
        'pbi': 'PBI-22', 'case_id': 'TC-22-01', 'scenario': 'Mapping Petugas ke Zona Wilayah', 'type': 'Positive', 
        'test_case': 'Assign petugas ke zona via Web', 'pre_cond': 'Akun admin',
        'steps': [
            (1, 'Login sebagai admin', 'Masuk ke Dashboard Admin.', 'Pass'),
            (2, 'Buka detail Zona Wilayah', 'Detail zona dan daftar petugas tampil.', 'Pass'),
            (3, 'Pilih petugas dan click Assign', 'Petugas berhasil di-assign ke zona.', 'Pass')
        ],
        'ket': '', 'evidence': 'SS_PBI-22_Mapping_Petugas.png'
    },
    {
        'pbi': 'PBI-23', 'case_id': 'TC-23-01', 'scenario': 'Validasi Wilayah Otomatis', 'type': 'Positive', 
        'test_case': 'Otomatisasi pencocokan koordinat', 'pre_cond': 'Akun masyarakat',
        'steps': [
            (1, 'Login sebagai masyarakat', 'Masuk ke Dashboard Masyarakat.', 'Pass'),
            (2, 'Klik tombol Buat Pengaduan', 'Form pengaduan dan peta penentuan lokasi tampil.', 'Pass'),
            (3, 'Pilih titik lokasi pada peta', 'Sistem memvalidasi koordinat dan mengisi field Zona secara otomatis.', 'Pass')
        ],
        'ket': '', 'evidence': 'SS_PBI-23_Validasi_Wilayah.png'
    }
]

row_start = 4

for tc in test_cases_data:
    first_step = True
    for step in tc['steps']:
        if first_step:
            ws.cell(row=row_start, column=1, value=tc['pbi'])
            ws.cell(row=row_start, column=2, value=tc['case_id'])
            ws.cell(row=row_start, column=3, value=tc['scenario'])
            ws.cell(row=row_start, column=4, value=tc['type'])
            ws.cell(row=row_start, column=5, value=tc['test_case'])
            ws.cell(row=row_start, column=6, value=tc['pre_cond'])
            ws.cell(row=row_start, column=11, value=tc['ket'])
            ws.cell(row=row_start, column=12, value=tc['evidence'])
            first_step = False
            
        ws.cell(row=row_start, column=7, value=step[0])
        ws.cell(row=row_start, column=8, value=step[1])
        ws.cell(row=row_start, column=9, value=step[2])
        ws.cell(row=row_start, column=10, value=step[3])
        
        row_start += 1

wb.save(output_path)
print('Excel file created successfully with multi-row steps.')
