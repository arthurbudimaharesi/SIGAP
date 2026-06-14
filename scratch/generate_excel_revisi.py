import pandas as pd
from openpyxl import load_workbook
import shutil
import os

os.makedirs('docs', exist_ok=True)
template_path = 'docs/Test Report RPL (1).xlsx'
output_path = 'docs/Test Report - Arthur (Final).xlsx'
shutil.copy(template_path, output_path)

wb = load_workbook(output_path)
ws = wb.active

test_cases_data = [
    {
        'pbi': 'PBI #1', 'case_id': 'TC.Login.001', 'scenario': 'Halaman Login', 'type': 'Positive', 
        'test_case': 'Login menggunakan kredensial valid', 'pre_cond': 'Akun pengguna terdaftar',
        'steps': [
            (1, 'Buka halaman login', 'Halaman login dengan form username dan password berhasil dimuat', 'Pass'),
            (2, 'Input email yang terdaftar (contoh: admin@example.com)', 'Field email terisi dengan format yang benar', 'Pass'),
            (3, 'Input password yang terdaftar', 'Field password terisi (disamarkan)', 'Pass'),
            (4, 'Klik tombol "Login"', 'Sistem memvalidasi kredensial dan mengarahkan ke halaman Dashboard sesuai role', 'Pass')
        ],
        'ket': 'Fitur berfungsi dengan baik sesuai ekspektasi', 'evidence': 'SS_Login.png'
    },
    {
        'pbi': 'PBI-01', 'case_id': 'TC-01-01', 'scenario': 'Administrasi Data Pelanggan', 'type': 'Positive', 
        'test_case': 'Proses CRUD Data Pelanggan', 'pre_cond': 'Login sebagai Admin',
        'steps': [
            (1, 'Arahkan kursor ke sidebar dan klik menu "Data Pelanggan"', 'Halaman daftar pelanggan (tabel) berhasil ditampilkan', 'Pass'),
            (2, 'Klik tombol "Tambah Pelanggan"', 'Modal / halaman form tambah pelanggan terbuka', 'Pass'),
            (3, 'Isi form pelanggan (Nama, Email, No. HP, Alamat) dengan data valid', 'Form terisi tanpa ada error validasi', 'Pass'),
            (4, 'Klik tombol "Simpan"', 'Muncul notifikasi sukses dan data pelanggan baru muncul di tabel', 'Pass'),
            (5, 'Klik tombol "Edit" pada pelanggan yang baru dibuat', 'Form edit terbuka dengan data pelanggan yang lama', 'Pass'),
            (6, 'Ubah data (misal: No. HP) dan klik "Update"', 'Muncul notifikasi sukses dan data di tabel ter-update', 'Pass'),
            (7, 'Klik tombol "Hapus" pada pelanggan tersebut', 'Muncul pop-up konfirmasi penghapusan', 'Pass'),
            (8, 'Klik "Ya, Hapus" pada pop-up konfirmasi', 'Notifikasi sukses muncul dan data hilang dari tabel', 'Pass')
        ],
        'ket': 'Fitur berfungsi dengan baik sesuai ekspektasi', 'evidence': 'SS_PBI-01_Pelanggan_Index.png'
    },
    {
        'pbi': 'PBI-02', 'case_id': 'TC-02-01', 'scenario': 'Administrasi Kategori Pengaduan', 'type': 'Positive', 
        'test_case': 'Proses CRUD Kategori Pengaduan', 'pre_cond': 'Login sebagai Admin',
        'steps': [
            (1, 'Klik menu "Kategori Pengaduan" di navigasi', 'Halaman tabel daftar kategori pengaduan ditampilkan', 'Pass'),
            (2, 'Klik tombol "Tambah Kategori"', 'Form tambah kategori (Nama Kategori, Deskripsi) muncul', 'Pass'),
            (3, 'Masukkan nama kategori "Infrastruktur" dan deskripsi terkait, lalu "Simpan"', 'Notifikasi sukses muncul, kategori "Infrastruktur" masuk ke tabel', 'Pass'),
            (4, 'Pilih tombol "Edit" pada baris kategori "Infrastruktur"', 'Form edit tampil dengan data sebelumnya', 'Pass'),
            (5, 'Ubah nama kategori menjadi "Infrastruktur & Jalan" dan "Update"', 'Data berhasil diubah dan tabel menampilkan nama baru', 'Pass'),
            (6, 'Klik icon "Hapus" untuk kategori tersebut dan konfirmasi', 'Kategori terhapus dari sistem dan tabel', 'Pass')
        ],
        'ket': 'Fitur berfungsi dengan baik sesuai ekspektasi', 'evidence': 'SS_PBI-02_Kategori_Index.png'
    },
    {
        'pbi': 'PBI-03 / PBI-20', 'case_id': 'TC-03-01', 'scenario': 'Administrasi Zona Wilayah', 'type': 'Positive', 
        'test_case': 'Proses Pembuatan dan Pengelolaan Zona Wilayah', 'pre_cond': 'Login sebagai Admin',
        'steps': [
            (1, 'Navigasi ke menu "Zona Wilayah"', 'Halaman tabel dan daftar Zona Wilayah ditampilkan', 'Pass'),
            (2, 'Klik "Tambah Zona Wilayah"', 'Form pengisian detail zona (Nama Zona, Kode Zona, Polygon Peta) tampil', 'Pass'),
            (3, 'Input data zona dan gambarkan poligon batas wilayah pada peta yang disediakan', 'Titik koordinat batas wilayah (polygon) terekam', 'Pass'),
            (4, 'Klik "Simpan"', 'Zona baru berhasil disimpan dan muncul di daftar tabel', 'Pass'),
            (5, 'Klik tombol "Edit" dan geser beberapa titik poligon pada peta', 'Perubahan batas koordinat terekam', 'Pass'),
            (6, 'Simpan perubahan', 'Batas wilayah ter-update dengan sukses', 'Pass')
        ],
        'ket': 'Fitur berfungsi dengan baik sesuai ekspektasi', 'evidence': 'SS_PBI-03_PBI-20_Zona_Index.png'
    },
    {
        'pbi': 'PBI-21', 'case_id': 'TC-21-01', 'scenario': 'Pengelolaan Zona Wilayah Pengaduan', 'type': 'Positive', 
        'test_case': 'Pemantauan Pengaduan Berdasarkan Zona oleh Supervisor', 'pre_cond': 'Login sebagai Supervisor',
        'steps': [
            (1, 'Klik menu "Pemantauan Zona" pada Dashboard Supervisor', 'Menampilkan daftar zona yang menjadi tanggung jawab supervisor', 'Pass'),
            (2, 'Pilih salah satu zona (misal: "Zona A")', 'Halaman memunculkan daftar pengaduan spesifik di Zona A', 'Pass'),
            (3, 'Gunakan filter status (Tertunda, Diproses, Selesai)', 'Daftar pengaduan tersaring sesuai status yang dipilih', 'Pass'),
            (4, 'Klik detail pada salah satu tiket pengaduan', 'Menampilkan informasi lengkap mengenai pelapor, lokasi, dan deskripsi masalah di zona tersebut', 'Pass')
        ],
        'ket': 'Fitur berfungsi dengan baik sesuai ekspektasi', 'evidence': 'SS_PBI-21_Supervisor_Zona.png'
    },
    {
        'pbi': 'PBI-22', 'case_id': 'TC-22-01', 'scenario': 'Mapping Petugas ke Zona Wilayah', 'type': 'Positive', 
        'test_case': 'Penugasan (Assign) Petugas Lapangan ke Zona', 'pre_cond': 'Login Admin, data Petugas & Zona tersedia',
        'steps': [
            (1, 'Buka menu "Mapping Petugas" atau detail Zona Wilayah', 'Halaman form mapping antara Zona dan Petugas tampil', 'Pass'),
            (2, 'Pilih "Zona A" pada dropdown Zona', 'Daftar petugas yang saat ini bertugas di Zona A ditampilkan (jika ada)', 'Pass'),
            (3, 'Cari dan pilih nama "Petugas 1" pada list Petugas yang tersedia', 'Petugas 1 terpilih dalam antrean penugasan', 'Pass'),
            (4, 'Klik tombol "Assign" atau "Tambahkan"', 'Muncul notifikasi sukses', 'Pass'),
            (5, 'Periksa daftar petugas di "Zona A"', '"Petugas 1" kini tercantum sebagai petugas aktif di Zona A', 'Pass')
        ],
        'ket': 'Fitur berfungsi dengan baik sesuai ekspektasi', 'evidence': 'SS_PBI-22_Mapping_Petugas.png'
    },
    {
        'pbi': 'PBI-23', 'case_id': 'TC-23-01', 'scenario': 'Validasi Wilayah Otomatis', 'type': 'Positive', 
        'test_case': 'Otomatisasi pencocokan koordinat dengan zona terdaftar', 'pre_cond': 'Login sebagai Masyarakat',
        'steps': [
            (1, 'Klik tombol "Buat Pengaduan" di Dashboard', 'Form pengaduan dan modul Peta Lokasi ditampilkan', 'Pass'),
            (2, 'Tarik (drag) pin lokasi pada peta ke area yang termasuk dalam "Zona B"', 'Sistem secara background memeriksa koordinat pin tersebut terhadap polygon zona', 'Pass'),
            (3, 'Perhatikan field "Zona Wilayah" pada form', 'Field "Zona Wilayah" otomatis terisi dengan "Zona B" tanpa input manual', 'Pass'),
            (4, 'Pindahkan pin lokasi ke area di luar cakupan zona yang ada', 'Sistem memvalidasi ulang koordinat', 'Pass'),
            (5, 'Perhatikan field "Zona Wilayah" pada form', 'Muncul peringatan "Lokasi di luar jangkauan layanan" atau field dikosongkan', 'Pass')
        ],
        'ket': 'Fitur berfungsi dengan baik sesuai ekspektasi', 'evidence': 'SS_PBI-23_Validasi_Wilayah.png'
    },
    {
        'pbi': 'PBI-43', 'case_id': 'TC-43-01', 'scenario': 'Integrasi GPS Lokasi', 'type': 'Positive', 
        'test_case': 'Mendapatkan lokasi otomatis menggunakan GPS', 'pre_cond': 'Login sebagai Masyarakat',
        'steps': [
            (1, 'Klik tombol "Buat Pengaduan"', 'Form pengaduan tampil', 'Pass'),
            (2, 'Klik icon target / tombol "Gunakan Lokasi Saat Ini"', 'Browser memunculkan prompt persetujuan akses lokasi (Location Permission)', 'Pass'),
            (3, 'Klik "Allow" (Izinkan) pada prompt browser', 'Sistem memanggil Geolocation API untuk mengambil koordinat pengguna saat ini', 'Pass'),
            (4, 'Tunggu proses mendapatkan koordinat', 'Pin pada peta langsung berpindah ke lokasi fisik pengguna secara akurat', 'Pass'),
            (5, 'Cek akurasi alamat', 'Alamat/koordinat yang muncul pada form sesuai dengan lokasi asli (real-time)', 'Pass')
        ],
        'ket': 'Fitur berfungsi dengan baik sesuai ekspektasi', 'evidence': 'SS_PBI-43_Integrasi_GPS.png'
    }
]

row_start = 4

for tc in test_cases_data:
    first_step = True
    for step in tc['steps']:
        if first_step:
            ws.cell(row=row_start, column=1, value=tc['pbi'])
            ws.cell(row=row_start, column=2, value=tc['case_id'])
            ws.cell(row=row_start, column=3, value=tc['scenario'])
            ws.cell(row=row_start, column=4, value=tc['type'])
            ws.cell(row=row_start, column=5, value=tc['test_case'])
            ws.cell(row=row_start, column=6, value=tc['pre_cond'])
            ws.cell(row=row_start, column=11, value=tc['ket'])
            ws.cell(row=row_start, column=12, value=tc['evidence'])
            first_step = False
            
        ws.cell(row=row_start, column=7, value=step[0])
        ws.cell(row=row_start, column=8, value=step[1])
        ws.cell(row=row_start, column=9, value=step[2])
        ws.cell(row=row_start, column=10, value=step[3])
        
        row_start += 1

wb.save(output_path)
print('Excel file (Revisi) created successfully.')
